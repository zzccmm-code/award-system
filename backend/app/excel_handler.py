"""Excel template generator and parser for award data import."""
import logging
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

SHEET_NAME = "\u5bfc\u5165\u6a21\u677f"
INSTR_SHEET = "\u586b\u5199\u8bf4\u660e"
FONT_NAME = "\u5fae\u8f6f\u96c5\u9ed1"

TEMPLATE_HEADERS = ["\u9879\u76ee\u540d\u79f0","\u83b7\u5956\u5e74\u5ea6","\u5956\u52b1\u7c7b\u578b","\u5956\u52b1\u7b49\u7ea7","\u5956\u52b1\u7ea7\u522b","\u5b8c\u6210\u5355\u4f4d","\u5b8c\u6210\u4eba","\u6388\u5956\u5355\u4f4d"]
AWARD_TYPE_OPTIONS = ["\u79d1\u6280\u8fdb\u6b65\u5956","\u6280\u672f\u53d1\u660e\u5956","\u81ea\u7136\u79d1\u5b66\u5956","\u4e13\u5229\u5956","\u6807\u51c6\u521b\u65b0\u5956","\u79d1\u6280\u6210\u679c\u8f6c\u5316\u5956"]
AWARD_LEVEL_OPTIONS = ["\u7279\u7b49\u5956","\u4e00\u7b49\u5956","\u4e8c\u7b49\u5956","\u4e09\u7b49\u5956","\u91d1\u5956","\u94f6\u5956"]

def generate_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    hf = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="1A3976", end_color="1A3976", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bd = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for ci, h in enumerate(TEMPLATE_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = bd
    wl = [40, 12, 14, 10, 12, 35, 25, 15]
    col_letters = [chr(64+i) for i in range(1, 9)]
    for i, w in enumerate(wl):
        ws.column_dimensions[col_letters[i]].width = w
    example = ["\u793a\u4f8b\u9879\u76ee", "2025", "\u79d1\u6280\u8fdb\u6b65\u5956", "\u4e00\u7b49\u5956", "\u56fd\u5bb6\u7ea7", "\u793a\u4f8b\u5355\u4f4d", "\u5f20\u4e09, \u674e\u56db", "\u56fd\u5bb6\u79d1\u6280\u90e8"]
    for ci, v in enumerate(example, 1):
        c = ws.cell(row=2, column=ci, value=v)
        c.font = Font(name=FONT_NAME, size=10, color="888888")
        c.alignment = Alignment(vertical="center"); c.border = bd
    ws2 = wb.create_sheet(INSTR_SHEET)
    instructions = [
        [INSTR_SHEET], [""],
        ["1. \u9879\u76ee\u540d\u79f0\uff1a\u5fc5\u586b\uff0c\u83b7\u5956\u9879\u76ee\u7684\u5168\u79f0"],
        ["2. \u83b7\u5956\u5e74\u5ea6\uff1a\u5fc5\u586b\uff0c\u56db\u4f4d\u6570\u5e74\u4efd\uff0c\u59822025"],
        ["3. \u5956\u52b1\u7c7b\u578b\uff1a\u5fc5\u586b\uff0c\u4e0b\u62c9\u9009\u62e9\u6216\u624b\u52a8\u8f93\u5165"],
        ["4. \u5956\u52b1\u7b49\u7ea7\uff1a\u5fc5\u586b\uff0c\u4e0b\u62c9\u9009\u62e9\u6216\u624b\u52a8\u8f93\u5165"],
        ["5. \u5b8c\u6210\u5355\u4f4d\uff1a\u5fc5\u586b\uff0c\u591a\u4e2a\u5355\u4f4d\u7528\u9017\u53f7\u5206\u9694"],
        ["6. \u5b8c\u6210\u4eba\uff1a\u5fc5\u586b\uff0c\u591a\u4eba\u7528\u9017\u53f7\u5206\u9694"],
        ["7. \u4fe1\u606f\u6765\u6e90\uff1a\u9009\u586b\uff0c\u9ed8\u8ba4\u4e3a\u624b\u52a8\u5f55\u5165"],
    ]
    for ri, rd in enumerate(instructions, 1):
        for ci, v in enumerate(rd, 1):
            ws2.cell(row=ri, column=ci, value=v).font = Font(name=FONT_NAME, size=11)
    out = BytesIO(); wb.save(out); out.seek(0)
    return out

def parse_excel(file_bytes: bytes) -> list[dict]:
    entries = []
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers, header_row_idx = [], 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        cleaned = [str(c).strip() if c else "" for c in row]
        if "\u9879\u76ee\u540d\u79f0" in cleaned:
            headers, header_row_idx = cleaned, row_idx; break
    if not headers:
        raise ValueError("Cannot find header row")
    cm = {}
    for i, h in enumerate(headers):
        if "\u9879\u76ee" in h: cm["name"] = i
        elif "\u5e74\u5ea6" in h: cm["year"] = i
        elif "\u5956\u52b1\u7c7b\u578b" in h: cm["type"] = i
        elif "\u7b49\u7ea7" in h: cm["level"] = i
        elif "\u5956\u52b1\u7ea7\u522b" in h or "\u7ea7\u522b" in h: cm["category"] = i
        elif "\u5b8c\u6210\u5355\u4f4d" in h: cm["unit"] = i
        elif "\u5b8c\u6210\u4eba" in h: cm["people"] = i
        elif "\u6388\u5956" in h or "\u6765\u6e90" in h: cm["source"] = i
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        cleaned = [str(c).strip() if c else "" for c in row]
        if not any(cleaned): continue
        name = cleaned[cm.get("name", 0)] if cm.get("name", 0) < len(cleaned) else ""
        if not name or name == "\u793a\u4f8b\u9879\u76ee" or len(name) < 3: continue
        try: year = int(cleaned[cm.get("year", 1)])
        except: year = 2025
        entries.append({
            "project_name": name,
            "award_year": year,
            "award_type": cleaned[cm.get("type", 2)] if cm.get("type", 2) < len(cleaned) else "\u79d1\u6280\u8fdb\u6b65\u5956",
            "award_level": cleaned[cm.get("level", 3)] if cm.get("level", 3) < len(cleaned) else "",
            "award_category": cleaned[cm.get("category", 4)] if cm.get("category", 4) < len(cleaned) else "",
            "completing_unit": cleaned[cm.get("unit", 5)] if cm.get("unit", 5) < len(cleaned) else "",
            "completers": cleaned[cm.get("people", 6)] if cm.get("people", 6) < len(cleaned) else "",
            "source": cleaned[cm.get("source", 7)] if cm.get("source", 7) < len(cleaned) else "\u624b\u52a8\u5f55\u5165",
        })
    return entries

def generate_export(items: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Awards"
    hf = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="1A3976", end_color="1A3976", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bd = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    headers = ["ID","\u9879\u76ee\u540d\u79f0","\u83b7\u5956\u5e74\u5ea6","\u5956\u52b1\u7c7b\u578b","\u5956\u52b1\u7b49\u7ea7","\u5b8c\u6210\u5355\u4f4d","\u5b8c\u6210\u4eba","\u6388\u5956\u5355\u4f4d"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = bd
    widths = [8, 40, 8, 14, 10, 35, 25, 18]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65+i)].width = w
    for ri, item in enumerate(items, 2):
        vals = [str(item.get(k,"")) for k in ["id","project_name","award_year","award_type","award_level","award_category","completing_unit","completers","source"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name=FONT_NAME, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = bd
    out = BytesIO(); wb.save(out); out.seek(0)
    return out
